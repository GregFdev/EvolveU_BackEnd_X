import pytest
import invoice_loader
from invoice_loader import Customer, Invoice, Product, \
    check_dupes, create_sheet_dict, \
        check_data_types, build_inv_list
# import openpyxl
from openpyxl import load_workbook

# should create worksheet within this test file using openpyxl

@pytest.fixture(params=["invoices_test.xlsx"])
def test_wb(request):
    wb = load_workbook(request.param)
    return wb 

@pytest.fixture(params=["customers", "invoices", "products"])
def test_ws(request):
    wb = load_workbook("invoices_test.xlsx")
    ws = wb[request.param]
    return ws

@pytest.fixture(params=[56])
def test_inv_num(request):
    return request.param

@pytest.fixture(params=["line_items"])
def test_lines(request):
    wb = load_workbook("invoices_test.xlsx")
    ws = wb[request.param]
    return ws

def test_instantiateclasses():
    objcust = Customer(1, "Greg Freson")
    objinv = Invoice(1, 1, "2019-01-01")
    objprod = Product(1, "Staples", 100)

    assert objcust.cust_name == "Greg Freson"
    assert objinv.invoice_num == 1
    assert objprod.prod_name == "Staples"

def test_invoice_addproduct():
    objinv = Invoice(1, 1, "2019-01-01")
    objprod = Product(1, "Staples", 100)
    objinv.add_product(objprod, 10)
    objprod = Product(2, "Hammer", 10)
    objinv.add_product(objprod, 20)
    objprod = Product(3, "Nails", 10)
    objinv.add_product(objprod, 5)
    objprod = Product(0,0,0)
    objinv.add_product(objprod, 5)

    assert objinv.invoice_line_items[1][0].prod_name == "Staples"
    assert objinv.invoice_line_items[2][1] == 20
    print('line items are: ', objinv.invoice_line_items)

    # objinv.add_product(1, 9)
    # assert objinv.invoice_line_items[1] == 19
    # objinv.add_product(2, 99)
    # assert objinv.invoice_line_items[2] == 99


def test_chkdupes(test_wb):
    assert check_dupes(test_wb['customers']) == [38]
    assert check_dupes(test_wb['invoices']) == []
    assert check_dupes(test_wb['products']) == [6]

# def test_blank_data_check(test_wb):
#     assert blank_data_check(loadws) == True

def test_check_data_types(test_wb):
    assert check_data_types(test_wb['customers']) != None
    assert check_data_types(test_wb['invoices']) == None
    assert check_data_types(test_wb['products']) == None


def test_createsheetdict(test_ws):
    print('dict is ', create_sheet_dict(test_ws))
    assert create_sheet_dict(test_ws) != {}

def test_build_inv_list(test_wb, test_inv_num):
    inv_dict = build_inv_list(test_wb)
    print(f'invoice dict is {inv_dict[test_inv_num].invoice_line_items}')
    assert inv_dict[test_inv_num].invoice_line_items[5][0].prod_name == 'wedge'
    assert len(inv_dict[test_inv_num].invoice_line_items) == 3
