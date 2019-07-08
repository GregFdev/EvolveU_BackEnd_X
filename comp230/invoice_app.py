from openpyxl import load_workbook
from invoice_loader import Customer, Product, Invoice

wb = load_workbook('invoices.xlsx')
print(wb.sheetnames)