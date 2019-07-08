import openpyxl
from openpyxl import load_workbook

class Customer():

    def __init__(self, cust_id, cust_name):
        self.cust_id = cust_id
        self.cust_name = cust_name

    def __repr__(self):
        return f"Customer name: {self.cust_name}"

class Product():

    def __init__(self, prod_id, prod_name, prod_cost):
        self.prod_id = prod_id
        self.prod_name = prod_name
        self.prod_cost = prod_cost
        
        
    def __repr__(self):
        return f"Product OBJ (Product Name: {self.prod_name})"

class Invoice():

    def __init__(self, invoice_num, cust_id, invoice_date):
        self.cust_id = cust_id
        self.invoice_num = invoice_num
        self.invoice_date = invoice_date
        self.invoice_line_items = {}  #  {prod_id: [prod_obj, qty]}
        # print('line items in constr', self.invoice_line_items)

    def __repr__(self):
        return "Invoice OBJ"

    def add_product(self, prod_obj, prod_qty):
        # print(f'invoice number {self.invoice_num}')
        if prod_obj != None and int(prod_qty) > 0:

            if prod_obj.prod_id not in self.invoice_line_items:
                # print('add product id num ', prod_obj.prod_id)
                self.invoice_line_items[prod_obj.prod_id] = [prod_obj, prod_qty]
            else:
                # print('add duplicate product id num ', prod_obj.prod_id)
                self.invoice_line_items[prod_obj.prod_id][1] += prod_qty
        else:
            print('no product or quantity to add')

def check_dupes(ws):

    seen = {}
    dupes = []

    for row in ws.iter_rows(min_row=2, values_only=True):

        if row[0] not in seen:
            seen[row[0]] = 1
        else:
            if seen[row[0]] == 1:
                dupes.append(row[0])
            seen[row[0]] += 1
    
    return dupes

def check_data_types(ws):  
    # checks if two lines have different data types
    # also catches lines with blanks
    xlist = []
    rowlist = []
    xrow = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        xrow += 1
        prevlist = rowlist
        rowlist = []
        for item in row:
            rowlist.append(type(item))
            print(type(item))
        print(rowlist)
        xlist.append(row)
    
        if (prevlist != []) and (rowlist != prevlist):
            return (f"Data types not consistent in row {row} at row num {xrow}")


def create_sheet_dict(ws):
    
    sheet_dict = {}

    # add validations here

    for row in ws.iter_rows(min_row=2, values_only=True):
        
        if ws.title == "customers":
            obj = Customer(row[0], row[1])
        elif ws.title == "invoices":
            obj = Invoice(row[0], row[1], row[2])
        elif ws.title == "products":
            obj = Product(row[0], row[1], row[2])
        else: 
            obj = None
        
        key = row[0]
        sheet_dict[key] = obj

    return sheet_dict

    
def build_inv_list(wb):
    inv_dict = create_sheet_dict(wb["invoices"])  # {invoice_num: inv_obj}
    prod_dict = create_sheet_dict(wb["products"]) # {prod_id: prod_obj}

    for value in wb["line_items"].iter_rows(min_row=2, values_only=True):
        #  each row is [inv#, prod_id, qty]
        invoice_num = value[0]
        prod_id = value[1]
        qty = value[2]

        prod_obj = prod_dict[prod_id]
        inv_dict[invoice_num].add_product(prod_obj, qty)  # add product for each row

    return inv_dict



