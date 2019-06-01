from openpyxl import load_workbook

wbmerge = load_workbook('invoices.xlsx') # new sheet to be merged into master
wbmaster = load_workbook('invoices_master.xlsx') # master sheet already existing

# first validate that the workbook to be merged into master is the correct format
# use master as the template

# check sheet names
if wbmerge.sheetnames != wbmaster.sheetnames:
    print('Incorrect Worksheet format')
else:
    print('correct sheet names')

#check each sheet header row
def headercheck(sheetname):
    ws_merge = wbmerge[sheetname]
    print('merge header is ', ws_merge[1][0].value)
    if wbmerge[sheetname] != wbmaster[sheetname]:
        print(f"Sheet {sheetname} header is not correct format")
        return False
    else:
        return True

print(headercheck("customers"))

def create_newmasterlist (ws_master, ws_merge):

    master_list = []
    merge_list = []

    for row in ws_master.iter_rows(min_row=2, values_only=True):
        master_list.append(row)

    for row in ws_merge.iter_rows(min_row=2, values_only=True):
        merge_list.append(row)

    master_list.extend(merge_list)
    print('length is now ', len(master_list))

    # convert to set then back to list to remove duplicates
    master_list = list(set(master_list))
    print('unique master length is now ', len(master_list))
    
    return master_list

def write_master(newmasterlist, sheet):
    numrows = len(newmasterlist)
    numcols = len(newmasterlist[0])
    # print(numcols, numrows)

    for c in range(1, numcols+1):
        for r in range(2, numrows+2):
            # print(c, r)
            sheet.cell(row=r, column=c, value=newmasterlist[r-2][c-1])

    wbmaster.save('invoices_master.xlsx')

xsheet = 'customers'

ws_new_master = create_newmasterlist(wbmaster[xsheet], wbmerge[xsheet])

print(ws_new_master)

write_master(ws_new_master, wbmaster[xsheet])



