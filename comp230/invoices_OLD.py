from openpyxl import load_workbook
wb = load_workbook('invoices.xlsx')
print(wb.sheetnames)

cust_ws = wb["customers"]
inv_ws = wb["invoices"]
lines_ws = wb["line_items"]
prod_ws = wb["products"]

print('cust ws is ', cust_ws)
invnum = int(input('Enter invoice Number to generate invoice: '))

for row in inv_ws.iter_rows(min_row=2, max_col=3, values_only=True):
    if row[0] == invnum:
        cust_ID = row[1]
        invoice_date = row[2]

for row in cust_ws.iter_rows(min_row=2, max_col=2, values_only=True):
    if row[0] == cust_ID:
        cust_name = row[1]

lines_list = []
tot_inv_cost = 0

for row in lines_ws.iter_rows(min_row=2, max_col=3, values_only=True):
    if row[0] == invnum:
        for rowprod in prod_ws.iter_rows(min_row=2, values_only=True):
            if rowprod[0] == row[1]:
                sub_total = row[2] * rowprod[2]
                lines_list.append({"prod_id": row[1], "prod_name": rowprod[1], "quantity": row[2], "prod_cost": rowprod[2], "subtotal": sub_total})
                tot_inv_cost += sub_total

invoice_output = load_workbook("invoice_template.xlsx")

ws = invoice_output["Invoice"]

ws['B3'] = invnum
ws["B4"] = invoice_date
ws["B5"] = cust_name
ws["B6"] = cust_ID
ws["E18"] = tot_inv_cost

for i, line in enumerate(lines_list):
    ws.cell(row=i+9, column=1, value=line["prod_name"])
    ws.cell(row=i+9, column=2, value=line["prod_id"])
    ws.cell(row=i+9, column=3, value=line["prod_cost"])
    ws.cell(row=i+9, column=4, value=line["quantity"])
    ws.cell(row=i+9, column=5, value=line["subtotal"])

newfilename = "invoice" + str(invnum) + ".xlsx"
invoice_output.save(newfilename)

